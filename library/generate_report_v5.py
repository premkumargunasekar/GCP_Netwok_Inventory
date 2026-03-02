#!/usr/bin/python

from ansible.module_utils.basic import AnsibleModule
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


def run_module():

    module_args = dict(
        inventory_data=dict(type='list', required=True),
        audit_logs=dict(type='list', required=True),
        project_env_map=dict(type='dict', required=True),
        report_dir=dict(type='str', required=True)
    )

    module = AnsibleModule(argument_spec=module_args)

    inventory_data = module.params['inventory_data']
    audit_logs = module.params['audit_logs']
    project_env_map = module.params['project_env_map']
    report_dir = module.params['report_dir']

    today = datetime.utcnow()
    today_str = today.strftime("%Y-%m-%d")
    seven_days_ago = today - timedelta(days=7)

    report_file = f"{report_dir}/gcp_network_governance_{today_str}.xlsx"

    green_fill = PatternFill(start_color="C6EFCE",
                             end_color="C6EFCE",
                             fill_type="solid")

    red_fill = PatternFill(start_color="FFC7CE",
                           end_color="FFC7CE",
                           fill_type="solid")

    header_font = Font(bold=True)

    wb = Workbook()

    # ==========================================================
    # 1️⃣ COMPONENT DETAILS SHEET
    # ==========================================================
    s1 = wb.active
    s1.title = "Components Details"

    headers = [
        "Environment",
        "Component Type",
        "Component Name",
        "Region",
        "Created Date",
        "Created By",
        "Status",
        "New Resource (Last 7 Days)"
    ]

    s1.append(headers)

    for col in range(1, len(headers) + 1):
        s1.cell(row=1, column=col).font = header_font

    # Map creators from audit logs
    creator_map = {}
    for log in audit_logs:
        proto = log.get("protoPayload", {})
        method = proto.get("methodName", "")
        resource_name = proto.get("resourceName", "")

        if method.endswith(".insert"):
            short_name = resource_name.split("/")[-1]
            principal = proto.get("authenticationInfo", {}).get("principalEmail", "NA")
            creator_map[short_name] = principal

    # Populate inventory
    for item in inventory_data:

        created_raw = item.get("creationTimestamp", "NA")
        created_date = created_raw.split("T")[0] if created_raw != "NA" else "NA"

        created_by = creator_map.get(item.get("name"), "NA")

        status = "Active"

        is_new = "NO"
        if created_raw != "NA":
            try:
                created_dt = datetime.strptime(
                    created_raw.replace("Z", "").split(".")[0],
                    "%Y-%m-%dT%H:%M:%S"
                )
                if created_dt >= seven_days_ago:
                    is_new = "YES"
            except Exception:
                pass

        s1.append([
            item.get("environment"),
            item.get("component_type"),
            item.get("name"),
            item.get("region", "global"),
            created_date,
            created_by,
            status,
            is_new
        ])

    # ==========================================================
    # 2️⃣ WEEKLY DATA SHEET
    # ==========================================================
    s2 = wb.create_sheet("Weekly_Data")

    weekly_headers = [
        "Environment",
        "Component Type",
        "Resource Name",
        "Region",
        "Principal Email",
        "Action Date",
        "Days Ago",
        "Status",
        "Action"
    ]

    s2.append(weekly_headers)

    for col in range(1, len(weekly_headers) + 1):
        s2.cell(row=1, column=col).font = header_font

    weekly_details = []
    processed = set()

    for log in audit_logs:

        proto = log.get("protoPayload", {})
        method = proto.get("methodName", "")
        resource_name = proto.get("resourceName", "")
        timestamp_str = log.get("timestamp")

        if not timestamp_str:
            continue

        timestamp_clean = timestamp_str.replace("Z", "").split(".")[0]
        log_time = datetime.strptime(timestamp_clean, "%Y-%m-%dT%H:%M:%S")

        if log_time < seven_days_ago:
            continue

        if method.endswith(".insert"):
            action = "Added"
        elif method.endswith(".delete"):
            action = "Deleted"
        else:
            continue

        unique_key = (resource_name, action, log_time.strftime("%Y-%m-%d"))
        if unique_key in processed:
            continue
        processed.add(unique_key)

        parts = resource_name.split("/")
        resource_short = parts[-1]

        region = "global"
        if "regions" in parts:
            region = parts[parts.index("regions") + 1]

        project = None
        if "projects" in parts:
            project = parts[parts.index("projects") + 1]

        env = project_env_map.get(project, "Unknown")

        if "networks" in method:
            comp_type = "VPC"
        elif "subnetworks" in method:
            comp_type = "Subnet"
        elif "routers" in method:
            comp_type = "Cloud Router"
        elif "forwardingRules" in method:
            comp_type = "Load Balancer"
        elif "interconnectAttachments" in method:
            comp_type = "Cloud Interconnect"
        else:
            comp_type = "Other"

        principal = proto.get("authenticationInfo", {}).get("principalEmail", "NA")

        days_ago = (today - log_time).days
        status = action

        weekly_details.append({
            "env": env,
            "region": region,
            "comp": comp_type,
            "action": action
        })

        s2.append([
            env,
            comp_type,
            resource_short,
            region,
            principal,
            log_time.strftime("%Y-%m-%d"),
            days_ago,
            status,
            action
        ])

        row_num = s2.max_row
        fill = green_fill if action == "Added" else red_fill

        for col in range(1, 10):
            s2.cell(row=row_num, column=col).fill = fill

    # ==========================================================
    # 3️⃣ EXECUTIVE SUMMARY SHEET
    # ==========================================================
    s3 = wb.create_sheet("Summary")

    summary_headers = [
        "Environment",
        "Region",
        "VPC's",
        "Cloud Router",
        "Load Balancers",
        "Cloud Interconnect",
        "Resource Addition Last 7 Days",
        "Resource Deletion Last 7 Days"
    ]

    s3.append(summary_headers)

    for col in range(1, len(summary_headers) + 1):
        s3.cell(row=1, column=col).font = header_font

    summary = {}

    # Count current inventory
    for item in inventory_data:
        key = (item.get("environment"), item.get("region", "global"))

        summary.setdefault(key, {
            "vpc": 0,
            "router": 0,
            "lb": 0,
            "interconnect": 0,
            "added_7": 0,
            "deleted_7": 0
        })

        comp = item.get("component_type")

        if comp == "VPC":
            summary[key]["vpc"] += 1
        elif comp == "Cloud Router":
            summary[key]["router"] += 1
        elif comp == "Load Balancer":
            summary[key]["lb"] += 1
        elif comp == "Cloud Interconnect":
            summary[key]["interconnect"] += 1

    # Count last 7 days changes
    for item in weekly_details:
        key = (item["env"], item["region"])

        if key not in summary:
            continue

        if item["action"] == "Added":
            summary[key]["added_7"] += 1
        else:
            summary[key]["deleted_7"] += 1

    for key, values in summary.items():
        env, region = key

        s3.append([
            env,
            region,
            values["vpc"],
            values["router"],
            values["lb"],
            values["interconnect"],
            values["added_7"],
            values["deleted_7"]
        ])

    wb.save(report_file)

    module.exit_json(
        changed=True,
        report_file=report_file
    )


def main():
    run_module()


if __name__ == '__main__':
    main()
